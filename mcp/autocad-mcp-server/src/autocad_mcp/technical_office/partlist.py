"""ERT partlist Excel generation from technical office QC reports."""

from __future__ import annotations

import json
import argparse
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from autocad_mcp.technical_office.job_metadata import JobMetadata, load_job_metadata
from autocad_mcp.technical_office.naming import safe_name


HEADERS = [
    "POZ NO",
    "CİNSİ",
    "GENİŞLİK",
    "UZUNLUK",
    "ADET",
    "KALİTE",
    "B.ALAN",
    "B.AĞIRLIK",
    "T.ALAN",
    "T.AĞIRLIK",
    "AÇIKLAMA",
]
SHEET_NAME = "Part_List_holes"


@dataclass
class PartlistRow:
    poz_no: str
    cinsi: str
    genislik: float | int
    uzunluk: float | int
    adet: int
    kalite: str
    birim_alan: float
    birim_agirlik: float
    aciklama: str


@dataclass
class PartlistResult:
    job_id: str
    project_name: str
    path: str | None
    rows: int
    manual_reviews: list[dict[str, Any]]
    ok: bool

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def create_partlist(job_dir: str | Path, output_dir: str | Path) -> PartlistResult:
    metadata = load_job_metadata(job_dir)
    output_root = Path(output_dir)
    rows, manual_reviews = build_partlist_rows(output_root)

    review_path = output_root / "partlist_manual_review_required.json"
    if manual_reviews:
        output_root.mkdir(parents=True, exist_ok=True)
        review_path.write_text(json.dumps(manual_reviews, indent=2, ensure_ascii=False), encoding="utf-8")
        return PartlistResult(
            job_id=metadata.job_id,
            project_name=metadata.project_name,
            path=None,
            rows=0,
            manual_reviews=manual_reviews,
            ok=False,
        )

    workbook_path = output_root / f"{safe_name(metadata.project_name)}_partlist.xlsx"
    write_partlist_excel(rows, metadata, workbook_path)
    review_path.unlink(missing_ok=True)
    return PartlistResult(
        job_id=metadata.job_id,
        project_name=metadata.project_name,
        path=str(workbook_path),
        rows=len(rows),
        manual_reviews=[],
        ok=True,
    )


def build_partlist_rows(output_dir: str | Path) -> tuple[list[PartlistRow], list[dict[str, Any]]]:
    output_root = Path(output_dir)
    summary_path = output_root / "job_summary.json"
    if not summary_path.exists():
        return [], [{"reason": "job_summary_missing", "detail": f"Missing job summary: {summary_path}"}]

    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    rows: list[PartlistRow] = []
    manual_reviews: list[dict[str, Any]] = []

    for review in summary.get("manual_reviews", []):
        manual_reviews.append(
            {
                "reason": "production_manual_review_pending",
                "poz_no": review.get("poz_no"),
                "source_pdf": review.get("source_pdf"),
                "detail": (
                    "Production has an unresolved manual review; the ERT partlist cannot be finalized "
                    "until the manager notification is handled."
                ),
                "source_reason": review.get("reason"),
            }
        )

    for produced in summary.get("produced", []):
        poz_no = str(produced.get("poz_no", "")).strip()
        if not produced.get("ok"):
            manual_reviews.append(
                {
                    "reason": "qc_not_ok",
                    "poz_no": poz_no or None,
                    "detail": "Only QC ok=true rows can enter the ERT partlist.",
                }
            )
            continue

        qc_path = _resolve_qc_path(output_root, produced)
        if not qc_path.exists():
            manual_reviews.append(
                {
                    "reason": "qc_missing",
                    "poz_no": poz_no or None,
                    "detail": f"Missing QC report: {qc_path}",
                }
            )
            continue

        row = _row_from_qc(json.loads(qc_path.read_text(encoding="utf-8")), qc_path)
        if isinstance(row, PartlistRow):
            rows.append(row)
        else:
            manual_reviews.append(row)

    if not rows and not manual_reviews:
        manual_reviews.append({"reason": "no_ok_qc_rows", "detail": "No produced QC rows were found."})

    return sorted(rows, key=_sort_key), manual_reviews


def write_partlist_excel(rows: list[PartlistRow], metadata: JobMetadata, path: str | Path) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)

    for row_index, row in enumerate(rows, start=2):
        sheet.append(
            [
                row.poz_no,
                row.cinsi,
                row.genislik,
                row.uzunluk,
                row.adet,
                row.kalite,
                row.birim_alan,
                row.birim_agirlik,
                f"=+G{row_index}*E{row_index}",
                f"=+H{row_index}*E{row_index}",
                row.aciklama,
            ]
        )

    _format_sheet(sheet)
    workbook.properties.title = f"{metadata.project_name} ERT Partlist"
    workbook.properties.creator = metadata.manager_agent
    workbook.save(output)
    return output


def _row_from_qc(qc: dict[str, Any], qc_path: Path) -> PartlistRow | dict[str, Any]:
    if not qc.get("ok") or qc.get("manual_review_required"):
        return {
            "reason": "qc_not_ok",
            "poz_no": qc.get("poz_no"),
            "detail": f"QC is not deliverable: {qc_path}",
        }

    spec = qc.get("plate_spec") or {}
    poz_no = str(spec.get("poz_no") or qc.get("poz_no") or "").strip()
    unit_surface = spec.get("unit_surface_area_m2")
    unit_weight = spec.get("unit_weight_kg")
    missing = [
        name
        for name, value in (
            ("unit_surface_area_m2", unit_surface),
            ("unit_weight_kg", unit_weight),
        )
        if value is None
    ]
    if missing:
        return {
            "reason": "partlist_metric_missing",
            "poz_no": poz_no or None,
            "detail": f"Missing {', '.join(missing)} in {qc_path}; ERT partlist values will not be guessed.",
        }

    return PartlistRow(
        poz_no=poz_no,
        cinsi=f"PL{_format_dimension(spec['thickness'])}",
        genislik=_whole_number(spec["height"]),
        uzunluk=_whole_number(spec["width"]),
        adet=int(spec.get("quantity") or 1),
        kalite=str(spec.get("material") or "UNKNOWN"),
        birim_alan=float(unit_surface),
        birim_agirlik=float(unit_weight),
        aciklama="Delikli" if spec.get("holes") or spec.get("slots") else "Deliksiz",
    )


def _resolve_qc_path(output_root: Path, produced: dict[str, Any]) -> Path:
    raw_path = Path(str(produced.get("qc_path") or ""))
    if raw_path.exists():
        return raw_path
    poz_no = str(produced.get("poz_no") or "")
    return output_root / safe_name(poz_no) / f"{safe_name(poz_no)}_qc.json"


def _format_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column, width in {
        "A": 14,
        "B": 14,
        "C": 12,
        "D": 12,
        "E": 10,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 12,
        "K": 14,
    }.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = "0.00"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _sort_key(row: PartlistRow) -> tuple[int, str]:
    cinsi = row.cinsi.upper()
    groups = [
        ("IPE",),
        ("HEA", "HEB", "HEM"),
        ("RHS", "SHS"),
        ("CHS",),
        ("UNP", "UPE"),
        ("L",),
        ("CC",),
        ("PL",),
    ]
    for index, prefixes in enumerate(groups):
        if any(cinsi.startswith(prefix) for prefix in prefixes):
            return index, row.poz_no
    return len(groups), row.poz_no


def _format_dimension(value: Any) -> str:
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:g}"


def _whole_number(value: Any) -> float | int:
    number = float(value)
    return int(number) if number.is_integer() else number


def main() -> None:
    parser = argparse.ArgumentParser(description="Create an ERT Part_List_holes Excel file from QC reports.")
    parser.add_argument("job_dir", help="Directory containing manager job.json metadata")
    parser.add_argument("output_dir", help="Technical office output directory containing job_summary.json")
    args = parser.parse_args()
    result = create_partlist(args.job_dir, args.output_dir)
    print(json.dumps(result.to_dict(), indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
