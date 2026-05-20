import csv
import json
from pathlib import Path

import ezdxf
import pytest
from openpyxl import load_workbook

from autocad_mcp.technical_office.dxf_writer import write_plate_dxf
from autocad_mcp.technical_office.approved_specs import load_approved_plate_specs
from autocad_mcp.technical_office.job_metadata import JobMetadataError, upsert_job_metadata
from autocad_mcp.technical_office.models import CornerReliefSpec, HoleSpec, PlateSpec
from autocad_mcp.technical_office.nc1_writer import write_plate_nc1
from autocad_mcp.technical_office.partlist import build_partlist_rows, create_partlist
from autocad_mcp.technical_office.pdf_reader import extract_pdf_content
from autocad_mcp.technical_office.pdf_reader import PdfExtraction, PdfPageContent
from autocad_mcp.technical_office.plate_extractor import build_plate_specs
from autocad_mcp.technical_office.pipeline import run_job
from autocad_mcp.technical_office.positions import PositionValidationError, load_positions_csv
from autocad_mcp.technical_office.qc import build_qc_report
from autocad_mcp.technical_office.relief_types import normalize_relief_type


def _workspace_root() -> Path:
    root = Path(__file__).resolve().parents[2]
    if root.name == "mcp":
        return root.parent
    return root


def _job_input(job_id: str, filename: str) -> Path:
    workspace = _workspace_root()
    suite_path = workspace / "workspace" / "imports" / "jobs" / job_id / filename
    if suite_path.exists():
        return suite_path
    return workspace / "data" / "imports" / "jobs" / job_id / filename


def test_all_agent_skill_references_exist():
    workspace = _workspace_root()
    missing = []
    for agent_file in (workspace / "agents").glob("*/AGENT.md"):
        text = agent_file.read_text(encoding="utf-8")
        for raw_path in _skill_paths_from_markdown(text):
            target = (agent_file.parent / raw_path).resolve()
            if not target.exists():
                missing.append(f"{agent_file.relative_to(workspace)} -> {raw_path}")
    assert missing == []


def test_active_technical_office_agents_reference_learning_and_partlist_skills():
    workspace = _workspace_root()
    expected = {
        "teknik-ofis-muduru": ["OGRENME_VE_HAFIZA_YONETIMI.md"],
        "autocad-uzman-1": ["OGRENME_VE_HAFIZA_YONETIMI.md"],
        "autocad-uzman-2": ["OGRENME_VE_HAFIZA_YONETIMI.md"],
        "dokuman-kontrol": ["OGRENME_VE_HAFIZA_YONETIMI.md", "ERT_PARTLIST_EXCEL_URETIMI.md"],
    }
    for agent, skill_names in expected.items():
        text = (workspace / "agents" / agent / "AGENT.md").read_text(encoding="utf-8")
        for skill_name in skill_names:
            assert skill_name in text


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("pah", "chamfer"),
        ("bevel", "chamfer"),
        ("chamfered", "chamfer"),
        ("round_relief", "round"),
        ("radius", "round"),
        ("polygon_contour", "polygon_contour"),
    ],
)
def test_relief_type_aliases_are_normalized_once(raw, expected):
    assert normalize_relief_type(raw) == expected


def test_polygon_contour_sentinel_is_not_loaded_as_corner_relief(tmp_path):
    job_dir = tmp_path / "job"
    job_dir.mkdir()
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "plates": [
                    {
                        "poz_no": "P100",
                        "width": 200,
                        "height": 100,
                        "thickness": 10,
                        "corner_reliefs": [{"type": "polygon_contour"}],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    approved = load_approved_plate_specs(job_dir)

    assert approved[0].spec.corner_reliefs == []


def test_positions_csv_validation_catches_missing_poz_no(tmp_path):
    path = tmp_path / "positions.csv"
    _write_csv(path, [{"poz_no": "", "quantity": "1"}])
    with pytest.raises(PositionValidationError, match="poz_no is required"):
        load_positions_csv(path)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("quantity", "-1", "quantity must be positive"),
        ("thickness_mm", "-10", "thickness_mm must be positive"),
        ("unit", "inch", "unsupported unit"),
    ],
)
def test_positions_csv_validation_catches_invalid_values(tmp_path, field, value, message):
    row = {"poz_no": "P1", "quantity": "1", "thickness_mm": "10"}
    row[field] = value
    path = tmp_path / "positions.csv"
    _write_csv(path, [row])
    with pytest.raises(PositionValidationError, match=message):
        load_positions_csv(path)


def test_plate_dxf_is_r2013_with_expected_layers_and_holes(tmp_path):
    spec = _sample_plate()
    path = tmp_path / "P100.dxf"
    write_plate_dxf(spec, path)

    doc = ezdxf.readfile(path)
    assert doc.dxfversion == "AC1027"
    assert "PLATE_OUTER" in doc.layers
    assert "PLATE_HOLES" in doc.layers

    entities = list(doc.modelspace())
    assert sum(1 for entity in entities if entity.dxftype() == "CIRCLE") == 2
    assert any(entity.dxftype() == "LWPOLYLINE" and entity.closed for entity in entities)


def test_plate_dxf_and_nc1_include_corner_reliefs(tmp_path):
    spec = PlateSpec(
        poz_no="P-CUGUL",
        width=658,
        height=192,
        thickness=12,
        material="S355J0",
        corner_reliefs=[
            CornerReliefSpec(corner="bottom_left", radius=25, relief_type="cugul"),
            CornerReliefSpec(corner="bottom_right", radius=25, relief_type="cugul"),
            CornerReliefSpec(corner="top_right", radius=25, relief_type="cugul"),
            CornerReliefSpec(corner="top_left", radius=25, relief_type="cugul"),
        ],
    )
    dxf_path = tmp_path / "P-CUGUL.dxf"
    nc1_path = tmp_path / "P-CUGUL.nc1"

    write_plate_dxf(spec, dxf_path)
    write_plate_nc1(spec, nc1_path)

    doc = ezdxf.readfile(dxf_path)
    outer = next(entity for entity in doc.modelspace() if entity.dxftype() == "LWPOLYLINE")
    points = outer.get_points("xyb")
    assert len(points) == 8
    assert sum(1 for _x, _y, bulge in points if abs(bulge) > 0.01) == 4
    assert all(bulge < -0.01 for _x, _y, bulge in points if abs(bulge) > 0.01)

    ak_points = [
        line
        for line in nc1_path.read_text(encoding="utf-8").splitlines()
        if line.startswith("  ") and len(line.split()) == 2
    ]
    assert len(ak_points) > 8


def test_round_corner_reliefs_remain_supported(tmp_path):
    spec = PlateSpec(
        poz_no="P-ROUND",
        width=200,
        height=100,
        thickness=10,
        corner_reliefs=[
            CornerReliefSpec(corner="bottom_left", radius=20, relief_type="round"),
            CornerReliefSpec(corner="bottom_right", radius=20, relief_type="round"),
            CornerReliefSpec(corner="top_right", radius=20, relief_type="round"),
            CornerReliefSpec(corner="top_left", radius=20, relief_type="round"),
        ],
    )
    dxf_path = tmp_path / "P-ROUND.dxf"

    write_plate_dxf(spec, dxf_path)

    doc = ezdxf.readfile(dxf_path)
    outer = next(entity for entity in doc.modelspace() if entity.dxftype() == "LWPOLYLINE")
    assert all(bulge > 0.01 for _x, _y, bulge in outer.get_points("xyb") if abs(bulge) > 0.01)


def test_chamfered_polygon_reliefs_are_qc_checked(tmp_path):
    spec = PlateSpec(
        poz_no="P-CHAMFER",
        width=240,
        height=150,
        thickness=8,
        corner_reliefs=[
            CornerReliefSpec(corner="bottom_left", radius=30, relief_type="chamfer"),
            CornerReliefSpec(corner="bottom_right", radius=30, relief_type="chamfer"),
            CornerReliefSpec(corner="top_right", radius=10, relief_type="chamfer"),
            CornerReliefSpec(corner="top_left", radius=10, relief_type="chamfer"),
        ],
    )
    dxf_path = tmp_path / "P-CHAMFER.dxf"
    nc1_path = tmp_path / "P-CHAMFER.nc1"

    write_plate_dxf(spec, dxf_path)
    write_plate_nc1(spec, nc1_path)
    qc = build_qc_report(spec, dxf_path, nc1_path)

    doc = ezdxf.readfile(dxf_path)
    outer = next(entity for entity in doc.modelspace() if entity.dxftype() == "LWPOLYLINE")
    points = outer.get_points("xyb")
    assert len(points) == 8
    assert all(abs(bulge) <= 1e-9 for _x, _y, bulge in points)
    assert qc["ok"] is True
    assert qc["dxf"]["polygon_corner_relief_count"] == 4


def test_asymmetric_chamfer_offsets_are_written_to_outer_contour(tmp_path):
    spec = PlateSpec(
        poz_no="P-ASYM-CHAMFER",
        width=240,
        height=150,
        thickness=8,
        corner_reliefs=[
            CornerReliefSpec(corner="top_left", radius=30, relief_type="chamfer", x_offset=30, y_offset=120),
            CornerReliefSpec(corner="top_right", radius=30, relief_type="chamfer", x_offset=30, y_offset=120),
        ],
    )
    dxf_path = tmp_path / "P-ASYM-CHAMFER.dxf"
    nc1_path = tmp_path / "P-ASYM-CHAMFER.nc1"

    write_plate_dxf(spec, dxf_path)
    write_plate_nc1(spec, nc1_path)
    qc = build_qc_report(spec, dxf_path, nc1_path)

    doc = ezdxf.readfile(dxf_path)
    outer = next(entity for entity in doc.modelspace() if entity.dxftype() == "LWPOLYLINE")
    points = [(round(x, 3), round(y, 3)) for x, y, _bulge in outer.get_points("xyb")]
    assert (240.0, 30.0) in points
    assert (210.0, 150.0) in points
    assert (30.0, 150.0) in points
    assert (0.0, 30.0) in points
    assert qc["ok"] is True
    assert qc["dxf"]["polygon_corner_relief_count"] == 2
    assert qc["nc1"]["ak_point_count"] == 6


def test_plate_nc1_snapshot(tmp_path):
    path = tmp_path / "P100.nc1"
    write_plate_nc1(_sample_plate(), path)
    assert path.read_text(encoding="utf-8") == (
        "ST\n"
        "  P100\n"
        "  S355\n"
        "  THICKNESS 10.000\n"
        "  QUANTITY 2\n"
        "AK\n"
        "  0.000 0.000\n"
        "  200.000 0.000\n"
        "  200.000 100.000\n"
        "  0.000 100.000\n"
        "BO\n"
        "  50.000 25.000 18.000\n"
        "  150.000 25.000 18.000\n"
        "EN\n"
    )


def test_qc_report_calculates_missing_partlist_metrics(tmp_path):
    spec = _sample_plate()
    dxf_path = tmp_path / "P100.dxf"
    nc1_path = tmp_path / "P100.nc1"
    write_plate_dxf(spec, dxf_path)
    write_plate_nc1(spec, nc1_path)

    qc = build_qc_report(spec, dxf_path, nc1_path)

    plate_spec = qc["plate_spec"]
    assert plate_spec["unit_surface_area_m2"] == pytest.approx(0.046)
    assert plate_spec["unit_weight_kg"] == pytest.approx(1.57)
    assert plate_spec["partlist_metrics_source"] == "calculated_geometry"
    assert set(plate_spec["partlist_metrics_calculated"]) == {"unit_surface_area_m2", "unit_weight_kg"}


def test_run_job_from_vector_pdf_and_manager_list(tmp_path):
    job_dir = tmp_path / "job-001"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "input.pdf",
        [
            "POZ: PDF-P100",
            "PLAKA 200x100x10 S355",
            "HOLE 50,25,D18",
            "HOLE 150,25,D18",
        ],
    )
    _write_csv(
        job_dir / "positions.csv",
        [
            {
                "poz_no": "MGR-P100",
                "page": "1",
                "quantity": "3",
                "thickness_mm": "12",
                "material": "S275",
            }
        ],
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is True
    assert len(result.produced) == 1
    produced = result.produced[0]
    assert produced.poz_no == "MGR-P100"
    assert Path(produced.dxf_path).name == "MGR-P100.dxf"
    assert Path(produced.nc1_path).name == "MGR-P100.nc1"

    qc = json.loads(Path(produced.qc_path).read_text(encoding="utf-8"))
    assert qc["ok"] is True
    assert qc["autocad_live_check"] == "skipped"
    assert qc["plate_spec"]["thickness"] == 12.0
    assert qc["plate_spec"]["material"] == "S275"


def test_run_job_autocad_live_autostart_failed_does_not_stop(tmp_path):
    job_dir = tmp_path / "job-live-fail"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "input.pdf",
        [
            "POZ: P-LIVE-FAIL",
            "PLAKA 120x60x8 S355",
        ],
    )

    result = run_job(
        job_dir,
        tmp_path / "out",
        autocad_live_policy="auto_start_if_needed",
        live_validator=lambda _path: "skipped_autostart_failed",
    )

    assert result.ok is True
    qc = json.loads(Path(result.produced[0].qc_path).read_text(encoding="utf-8"))
    assert qc["autocad_live_check"] == "skipped_autostart_failed"
    assert qc["ok"] is True


def test_run_job_autocad_live_ok_is_recorded(tmp_path):
    job_dir = tmp_path / "job-live-ok"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "input.pdf",
        [
            "POZ: P-LIVE-OK",
            "PLAKA 120x60x8 S355",
        ],
    )

    result = run_job(
        job_dir,
        tmp_path / "out",
        autocad_live_policy="auto_start_if_needed",
        live_validator=lambda _path: "ok",
    )

    assert result.ok is True
    qc = json.loads(Path(result.produced[0].qc_path).read_text(encoding="utf-8"))
    assert qc["autocad_live_check"] == "ok"


def test_run_job_accepts_single_pdf_when_input_pdf_is_missing(tmp_path):
    job_dir = tmp_path / "job-single-pdf"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "1001.pdf",
        [
            "POZ: P1001",
            "PLAKA 90x40x6 S235",
        ],
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is True
    assert result.produced[0].poz_no == "P1001"


def test_run_job_processes_all_pdfs_when_input_pdf_is_missing(tmp_path):
    job_dir = tmp_path / "job-multi-pdf"
    job_dir.mkdir()
    _write_vector_pdf(job_dir / "1001.pdf", ["POZ: P1001", "PLAKA 90x40x6 S235"])
    _write_vector_pdf(job_dir / "1142.pdf", ["POZ: P1142", "PLAKA 120x50x8 S355"])

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is True
    assert sorted(item.poz_no for item in result.produced) == ["P1001", "P1142"]
    assert (tmp_path / "out" / "P1001" / "P1001.dxf").exists()
    assert (tmp_path / "out" / "P1142" / "P1142.nc1").exists()


def test_run_job_can_select_one_pdf_from_multi_pdf_job(tmp_path):
    job_dir = tmp_path / "job-select-pdf"
    job_dir.mkdir()
    _write_vector_pdf(job_dir / "1001.pdf", ["POZ: P1001", "PLAKA 90x40x6 S235"])
    _write_vector_pdf(job_dir / "1142.pdf", ["POZ: P1142", "PLAKA 120x50x8 S355"])

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off", input_pdf="1142.pdf")

    assert result.ok is True
    assert [item.poz_no for item in result.produced] == ["P1142"]
    assert not (tmp_path / "out" / "P1001").exists()


def test_run_job_applies_manager_page_exclusions(tmp_path):
    job_dir = tmp_path / "job-page-exclusion"
    job_dir.mkdir()
    _write_multi_page_vector_pdf(
        job_dir / "input.pdf",
        [
            ["POZ: COVER1", "PROJECT TITLE PAGE"],
            ["POZ: P100", "PLAKA 90x40x6 S235"],
        ],
    )
    (job_dir / "page_exclusions.json").write_text(
        json.dumps(
            {
                "excluded_pages": [
                    {
                        "page": 1,
                        "reason": "manager_confirmed_non_plate_page",
                        "note": "title page",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, output_root=tmp_path / "out", autocad_live_policy="off")

    assert result.manual_reviews == []
    assert [item.poz_no for item in result.produced] == ["P100"]
    applied = json.loads((tmp_path / "out" / "page_exclusions_applied.json").read_text(encoding="utf-8"))
    assert applied["excluded_pages"][0]["page"] == 1
    candidates = json.loads((tmp_path / "out" / "extraction_candidates.json").read_text(encoding="utf-8"))
    assert [candidate["page"] for candidate in candidates["candidates"]] == [2]


def test_run_job_reads_plate_profile_table_without_prefixed_poz(tmp_path):
    job_dir = tmp_path / "job-profile-table"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "1001.pdf",
        [
            "1001",
            "PL25*160",
            "S355J0",
            "676",
            "2",
            "0.26",
            "21.23",
            "Part / Assembly",
            "Profile",
            "Material",
            "Q.ty",
            "Length [mm]",
            "Scale",
            "1001",
            "Object",
        ],
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is True
    qc = json.loads(Path(result.produced[0].qc_path).read_text(encoding="utf-8"))
    assert result.produced[0].poz_no == "1001"
    assert qc["plate_spec"]["width"] == 676.0
    assert qc["plate_spec"]["height"] == 160.0
    assert qc["plate_spec"]["thickness"] == 25.0
    assert qc["plate_spec"]["material"] == "S355J0"
    assert qc["plate_spec"]["quantity"] == 2
    assert qc["plate_spec"]["unit_surface_area_m2"] == 0.26
    assert qc["plate_spec"]["unit_weight_kg"] == 21.23


def test_partlist_requires_job_metadata(tmp_path):
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    (output_dir / "job_summary.json").write_text(
        json.dumps({"produced": []}, indent=2),
        encoding="utf-8",
    )

    with pytest.raises(JobMetadataError, match="Missing job metadata"):
        create_partlist(tmp_path / "job-missing-metadata", output_dir)


def test_partlist_requires_project_name_in_job_metadata(tmp_path):
    job_dir = tmp_path / "job-empty-project"
    job_dir.mkdir()
    (job_dir / "job.json").write_text(
        json.dumps({"job_id": "job-empty-project", "project_name": ""}, indent=2),
        encoding="utf-8",
    )
    output_dir = tmp_path / "out"
    output_dir.mkdir()
    (output_dir / "job_summary.json").write_text(
        json.dumps({"produced": []}, indent=2),
        encoding="utf-8",
    )

    with pytest.raises(JobMetadataError, match="project_name is required"):
        create_partlist(job_dir, output_dir)


def test_partlist_row_from_qc_matches_ert_format(tmp_path):
    job_dir = tmp_path / "job-partlist"
    job_dir.mkdir()
    upsert_job_metadata(job_dir, "EAF Steel Platform")
    output_dir = tmp_path / "out"
    plate_dir = output_dir / "1001"
    plate_dir.mkdir(parents=True)
    qc_path = plate_dir / "1001_qc.json"
    qc_path.write_text(
        json.dumps(
            {
                "poz_no": "1001",
                "ok": True,
                "manual_review_required": False,
                "plate_spec": {
                    "poz_no": "1001",
                    "width": 676.0,
                    "height": 160.0,
                    "thickness": 25.0,
                    "material": "S355J0",
                    "quantity": 2,
                    "unit_surface_area_m2": 0.26,
                    "unit_weight_kg": 21.23,
                    "holes": [{"x": 78.0, "y": 35.0, "diameter": 22.0}],
                    "slots": [],
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (output_dir / "job_summary.json").write_text(
        json.dumps(
            {
                "produced": [
                    {
                        "poz_no": "1001",
                        "qc_path": str(qc_path),
                        "ok": True,
                    }
                ]
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = create_partlist(job_dir, output_dir)

    assert result.ok is True
    assert result.rows == 1
    workbook = load_workbook(result.path, data_only=False)
    sheet = workbook["Part_List_holes"]
    assert [sheet.cell(row=1, column=index).value for index in range(1, 12)] == [
        "POZ NO",
        "CİNSİ",
        "GENİŞLİK",
        "UZUNLUK",
        "ADET",
        "KALİTE",
        "B.ALAN",
        "B.AĞIRLIK",
        "T.ALAN",
        "T.AĞIRLIK",
        "AÇIKLAMA",
    ]
    assert [sheet.cell(row=2, column=index).value for index in range(1, 12)] == [
        "1001",
        "PL25",
        160,
        676,
        2,
        "S355J0",
        0.26,
        21.23,
        "=+G2*E2",
        "=+H2*E2",
        "Delikli",
    ]
    assert Path(result.path).name == "EAF_Steel_Platform_partlist.xlsx"


def test_partlist_missing_metrics_are_calculated_from_geometry(tmp_path):
    output_dir = tmp_path / "out"
    plate_dir = output_dir / "P100"
    plate_dir.mkdir(parents=True)
    qc_path = plate_dir / "P100_qc.json"
    qc_path.write_text(
        json.dumps(
            {
                "poz_no": "P100",
                "ok": True,
                "manual_review_required": False,
                "plate_spec": {
                    "poz_no": "P100",
                    "width": 200.0,
                    "height": 100.0,
                    "thickness": 10.0,
                    "material": "S355",
                    "quantity": 1,
                    "holes": [],
                    "slots": [],
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (output_dir / "job_summary.json").write_text(
        json.dumps({"produced": [{"poz_no": "P100", "qc_path": str(qc_path), "ok": True}]}, indent=2),
        encoding="utf-8",
    )

    rows, manual_reviews = build_partlist_rows(output_dir)

    assert manual_reviews == []
    assert len(rows) == 1
    assert rows[0].birim_alan == pytest.approx(0.046)
    assert rows[0].birim_agirlik == pytest.approx(1.57)


def test_partlist_blocks_when_production_manual_review_is_pending(tmp_path):
    job_dir = tmp_path / "job-partlist-blocked"
    job_dir.mkdir()
    upsert_job_metadata(job_dir, "EAF Steel Platform")
    output_dir = tmp_path / "out"
    plate_dir = output_dir / "1001"
    plate_dir.mkdir(parents=True)
    qc_path = plate_dir / "1001_qc.json"
    qc_path.write_text(
        json.dumps(
            {
                "poz_no": "1001",
                "ok": True,
                "manual_review_required": False,
                "plate_spec": {
                    "poz_no": "1001",
                    "width": 200.0,
                    "height": 100.0,
                    "thickness": 10.0,
                    "material": "S355",
                    "quantity": 1,
                    "unit_surface_area_m2": 0.02,
                    "unit_weight_kg": 1.57,
                    "holes": [],
                    "slots": [],
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    (output_dir / "job_summary.json").write_text(
        json.dumps(
            {
                "produced": [{"poz_no": "1001", "qc_path": str(qc_path), "ok": True}],
                "manual_reviews": [
                    {
                        "reason": "hole_geometry_not_found",
                        "poz_no": "1006",
                        "source_pdf": "1701.pdf",
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = create_partlist(job_dir, output_dir)

    assert result.ok is False
    assert result.path is None
    assert result.manual_reviews[0]["reason"] == "production_manual_review_pending"
    assert result.manual_reviews[0]["poz_no"] == "1006"
    assert (output_dir / "partlist_manual_review_required.json").exists()


def test_run_job_can_create_job_metadata_with_project_name(tmp_path):
    job_dir = tmp_path / "job-project-name"
    job_dir.mkdir()
    _write_vector_pdf(job_dir / "input.pdf", ["POZ: P100", "PLAKA 90x40x6 S235"])

    run_job(job_dir, tmp_path / "out", autocad_live_policy="off", project_name="Demo Project")

    metadata = json.loads((job_dir / "job.json").read_text(encoding="utf-8"))
    assert metadata["job_id"] == "job-project-name"
    assert metadata["project_name"] == "Demo Project"


def test_run_job_and_partlist_e2e_from_manager_project_name(tmp_path):
    job_dir = tmp_path / "job-ert-e2e"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "1001.pdf",
        [
            "1001",
            "PL25*160",
            "S355J0",
            "676",
            "2",
            "0.26",
            "21.23",
            "Part / Assembly",
            "Scale",
            "1001",
            "Object",
        ],
    )
    _write_vector_pdf(
        job_dir / "1142.pdf",
        [
            "1142",
            "PL20*400",
            "S355J0",
            "5200",
            "1",
            "4.38",
            "326.56",
            "Part / Assembly",
            "Scale",
            "1142",
            "Object",
        ],
    )

    result = run_job(
        job_dir,
        tmp_path / "out",
        autocad_live_policy="off",
        project_name="EAF Steel Platform",
    )
    partlist = create_partlist(job_dir, tmp_path / "out")

    assert result.ok is True
    assert partlist.ok is True
    workbook = load_workbook(partlist.path, data_only=False)
    sheet = workbook["Part_List_holes"]
    assert sheet.max_row == 3
    assert [sheet.cell(row=row, column=1).value for row in (2, 3)] == ["1001", "1142"]
    assert [sheet.cell(row=row, column=2).value for row in (2, 3)] == ["PL25", "PL20"]
    assert [sheet.cell(row=row, column=9).value for row in (2, 3)] == ["=+G2*E2", "=+G3*E3"]


def test_run_job_requires_review_when_hole_callout_has_no_coordinates(tmp_path):
    job_dir = tmp_path / "job-profile-table-holes"
    job_dir.mkdir()
    _write_vector_pdf(
        job_dir / "1001.pdf",
        [
            "1001",
            "PL25*160",
            "S355J0",
            "676",
            "2",
            "0.26",
            "21.23",
            "Part / Assembly",
            "Profile",
            "Material",
            "Q.ty",
            "Length [mm]",
            "Scale",
            "1001",
            "Object",
            "14*Ø22",
        ],
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is False
    assert result.produced == []
    assert result.manual_reviews[0]["reason"] == "hole_geometry_not_found"
    assert result.manual_reviews[0]["poz_no"] == "1001"
    notifications = json.loads((tmp_path / "out" / "manager_notifications.json").read_text(encoding="utf-8"))
    assert notifications[0]["role"] == "teknik-ofis-muduru"
    assert notifications[0]["poz_no"] == "1001"


def test_mark_column_bottom_table_supplies_poz_for_manual_review():
    extraction = PdfExtraction(
        pages=[
            PdfPageContent(
                page_number=4,
                text=(
                    "FRONT VIE\nTOP VIE\n402\nPL1022\nS55JR\n5.5\n0.1\n5.\n0.\n5.\n"
                    "DANIELI CONSTRUCTION\nMARK\nPROFILE\nMATERIAL\n.T\nLENGTH mm\nAREA m2\n"
                    "EIGHT kg\nTOTAL\nIN ASSEMBLIES\n"
                ),
                vector_operator_count=50,
                raw_length=100,
            )
        ],
        manual_review_required=False,
        confidence=0.95,
        notes=[],
    )

    result = build_plate_specs(extraction)

    assert result.plates == []
    assert result.manual_reviews[0].reason == "plate_geometry_not_found"
    assert result.manual_reviews[0].poz_no == "402"


def test_vector_holes_are_mapped_from_dimension_chain():
    circles = []
    for y in (126.48, 177.48):
        for x in (278.88, 318.60, 358.26, 397.92, 437.64, 477.36, 516.96):
            circles.append({"cx_pt": x, "cy_pt": y, "diameter_pt": 12.48})
    extraction = PdfExtraction(
        pages=[
            PdfPageContent(
                page_number=1,
                text=(
                    "1001\nPL25*160\nS355J0\n676\n2\n0.26\n21.23\n"
                    "Part / Assembly\nScale\n1001\nObject\n14*Ø22\n"
                    "70\n178\n78\n676\n70\n70\n70\n70\n70\n"
                ),
                vector_operator_count=100,
                raw_length=100,
                vector_circles=circles,
            )
        ],
        manual_review_required=False,
        confidence=0.95,
        notes=[],
    )

    result = build_plate_specs(extraction)

    assert result.manual_reviews == []
    spec = result.plates[0]
    assert spec.poz_no == "1001"
    assert len(spec.holes) == 14
    assert [hole.x for hole in spec.holes[:7]] == [78, 148, 218, 288, 358, 428, 498]
    assert {hole.y for hole in spec.holes} == {35.0, 125.0}
    assert {hole.diameter for hole in spec.holes} == {22.0}


def test_1701_position_1004_extracts_r25_corner_reliefs():
    pdf_path = _job_input("test-001", "1701.pdf")
    extraction = extract_pdf_content(pdf_path)

    result = build_plate_specs(extraction)

    spec = next(plate for plate in result.plates if plate.poz_no == "1004")
    assert spec.width == 658.0
    assert spec.height == 192.0
    assert sorted((relief.corner, relief.radius, relief.relief_type) for relief in spec.corner_reliefs) == [
        ("bottom_left", 25.0, "cugul"),
        ("bottom_right", 25.0, "cugul"),
        ("top_left", 25.0, "cugul"),
        ("top_right", 25.0, "cugul"),
    ]


def test_1701_position_1006_extracts_two_d13_holes():
    pdf_path = _job_input("test-001", "1701.pdf")
    extraction = extract_pdf_content(pdf_path)

    result = build_plate_specs(extraction)

    spec = next(plate for plate in result.plates if plate.poz_no == "1006")
    assert spec.width == 185.0
    assert spec.height == 100.0
    assert spec.thickness == 15.0
    assert spec.material == "S355J0"
    assert spec.quantity == 1
    assert spec.unit_surface_area_m2 == 0.04
    assert spec.unit_weight_kg == 2.03
    assert [(hole.x, hole.y, hole.diameter) for hole in spec.holes] == [
        (35.0, 65.0, 13.0),
        (150.0, 65.0, 13.0),
    ]
    assert all(review.poz_no != "1006" for review in result.manual_reviews)


def test_1142_extracts_compact_metrics_and_multi_diameter_holes():
    pdf_path = _job_input("test-001", "1142.pdf")
    extraction = extract_pdf_content(pdf_path)

    result = build_plate_specs(extraction)

    assert result.manual_reviews == []
    spec = result.plates[0]
    assert spec.poz_no == "1142"
    assert spec.width == 5200.0
    assert spec.height == 400.0
    assert spec.thickness == 20.0
    assert spec.unit_surface_area_m2 == 4.38
    assert spec.unit_weight_kg == 326.56
    assert len(spec.holes) == 13
    assert [(hole.x, hole.y, hole.diameter) for hole in spec.holes[:4]] == [
        (35.0, 342.0, 13.0),
        (45.0, 260.0, 18.0),
        (115.0, 260.0, 18.0),
        (4898.0, 250.0, 13.0),
    ]
    assert [(hole.x, hole.y, hole.diameter) for hole in spec.holes[-4:]] == [
        (491.5, 35.0, 13.0),
        (941.5, 35.0, 13.0),
        (1391.5, 35.0, 13.0),
        (1841.5, 35.0, 13.0),
    ]


def test_run_job_marks_scanned_or_empty_pdf_for_manual_review(tmp_path):
    job_dir = tmp_path / "job-scan"
    job_dir.mkdir()
    (job_dir / "input.pdf").write_bytes(b"%PDF-1.4\n% no readable streams\n")

    result = run_job(job_dir, tmp_path / "out")

    assert result.ok is False
    assert result.produced == []
    assert result.manual_reviews[0]["reason"] == "manual_review_required"
    assert (tmp_path / "out" / "manual_review_required.json").exists()


def test_run_job_groups_vector_pdf_without_text_as_visual_review_candidate(tmp_path):
    job_dir = tmp_path / "job-visual"
    job_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf")

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is False
    assert result.produced == []
    assert len(result.manual_reviews) == 1
    assert result.manual_reviews[0]["reason"] == "visual_text_required"
    assert "sayfa numarasi" in result.manual_reviews[0]["detail"]
    assert "kullaniciya kurulum talimati degil" in result.manual_reviews[0]["detail"]
    assert "Teknik ofis muduru" in result.manual_reviews[0]["next_action"]

    diagnostics = json.loads((tmp_path / "out" / "pdf_diagnostics.json").read_text(encoding="utf-8"))
    assert diagnostics["pdfs"][0]["classification"] == "visual_text_required"
    assert diagnostics["summary"]["approval_required"] is True

    candidates = json.loads((tmp_path / "out" / "extraction_candidates.json").read_text(encoding="utf-8"))
    assert candidates["approval_required"] is True
    assert candidates["candidates"][0]["approval_required"] is True

    notifications = json.loads((tmp_path / "out" / "manager_notifications.json").read_text(encoding="utf-8"))
    assert "was not produced" not in notifications[0]["message"]
    assert notifications[0]["approval_required"] is True
    assert "kullaniciya kurulum talimati" not in notifications[0]["next_action"]
    assert "Teknik ofis muduru" in notifications[0]["next_action"]


def test_run_job_uses_manager_approved_specs_after_visual_review(tmp_path):
    job_dir = tmp_path / "job-approved"
    job_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf")
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "approved_by": "teknik-ofis-muduru",
                "plates": [
                    {
                        "poz_no": "P100",
                        "width": 200,
                        "height": 100,
                        "thickness": 10,
                        "material": "S355",
                        "quantity": 1,
                        "source_pdf": "input.pdf",
                        "source_page": 1,
                        "holes": [{"x": 50, "y": 25, "diameter": 18}],
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is True
    assert result.manual_reviews == []
    assert [item.poz_no for item in result.produced] == ["P100"]
    assert (tmp_path / "out" / "P100" / "P100.dxf").exists()
    qc = json.loads((tmp_path / "out" / "P100" / "P100_qc.json").read_text(encoding="utf-8"))
    assert qc["ok"] is True
    assert qc["source_pdf"] == "input.pdf"
    assert (tmp_path / "out" / "pdf_diagnostics.json").exists()
    assert (tmp_path / "out" / "extraction_candidates.json").exists()


def test_run_job_blocks_manager_approved_specs_when_visual_pages_are_missing(tmp_path):
    job_dir = tmp_path / "job-approved-missing-pages"
    job_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf", page_count=4)
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "approved_by": "teknik-ofis-muduru",
                "plates": [
                    {
                        "poz_no": "P100",
                        "width": 200,
                        "height": 100,
                        "thickness": 10,
                        "material": "S355",
                        "quantity": 1,
                        "source_pdf": "input.pdf",
                        "source_page": 1,
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, tmp_path / "out", autocad_live_policy="off")

    assert result.ok is False
    assert [item.poz_no for item in result.produced] == ["P100"]
    assert result.produced[0].ok is True
    assert result.manual_reviews[0]["reason"] == "approved_specs_missing_visual_pages"
    assert "Eksik sayfalar: 2-4" in result.manual_reviews[0]["detail"]
    assert (tmp_path / "out" / "manual_review_required.json").exists()


def test_run_job_blocks_open_manager_geometry_notes(tmp_path):
    job_dir = tmp_path / "job-approved-geometry-note"
    output_dir = tmp_path / "out"
    job_dir.mkdir()
    output_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf")
    (output_dir / "manager_issue_notes.jsonl").write_text(
        json.dumps(
            {
                "status": "open",
                "tags": ["pah/kose eksigi", "poligon kontur"],
                "affected_pozs": ["206"],
            }
        )
        + "\n",
        encoding="utf-8",
    )
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "approved_by": "teknik-ofis-muduru",
                "plates": [
                    {
                        "poz_no": "206",
                        "width": 240,
                        "height": 150,
                        "thickness": 8,
                        "material": "S275J2",
                        "quantity": 10,
                        "source_pdf": "input.pdf",
                        "source_page": 1,
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, output_dir, autocad_live_policy="off")

    assert result.ok is False
    assert result.produced[0].ok is False
    assert result.manual_reviews[0]["reason"] == "manager_geometry_issue_open"
    assert result.manual_reviews[0]["poz_no"] == "206"
    qc = json.loads((output_dir / "206" / "206_qc.json").read_text(encoding="utf-8"))
    assert qc["ok"] is False
    assert qc["manual_review_required"] is True


def test_run_job_accepts_single_pdf_alias_when_geometry_note_is_resolved(tmp_path):
    job_dir = tmp_path / "job-approved-resolved-geometry-note"
    output_dir = tmp_path / "out"
    job_dir.mkdir()
    output_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf")
    (output_dir / "manager_issue_notes.jsonl").write_text(
        json.dumps(
            {
                "status": "open",
                "tags": ["pah/kose eksigi", "poligon kontur"],
                "affected_pozs": ["206"],
            }
        )
        + "\n",
        encoding="utf-8",
    )
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "approved_by": "teknik-ofis-muduru",
                "plates": [
                    {
                        "poz_no": "206",
                        "width": 240,
                        "height": 150,
                        "thickness": 8,
                        "material": "S275J2",
                        "quantity": 10,
                        "source_pdf": "stale-candidate-name.pdf",
                        "source_page": 1,
                        "corner_reliefs": [
                            {
                                "corner": "top_left",
                                "relief_type": "chamfer",
                                "radius": 30,
                                "x_offset": 30,
                                "y_offset": 120,
                            }
                        ],
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, output_dir, autocad_live_policy="off")

    assert result.ok is True
    assert result.manual_reviews == []
    qc = json.loads((output_dir / "206" / "206_qc.json").read_text(encoding="utf-8"))
    assert qc["ok"] is True
    assert qc["manual_review_required"] is False
    assert qc["source_pdf"] == "input.pdf"


def test_run_job_preserves_approved_polygon_vertices_for_qc(tmp_path):
    job_dir = tmp_path / "job-approved-polygon"
    output_dir = tmp_path / "out"
    job_dir.mkdir()
    _write_vector_only_pdf(job_dir / "input.pdf")
    (job_dir / "approved_plate_specs.json").write_text(
        json.dumps(
            {
                "approved_by": "teknik-ofis-muduru",
                "plates": [
                    {
                        "poz_no": "4042",
                        "width": 156.5,
                        "height": 175,
                        "thickness": 10,
                        "material": "S355JR",
                        "quantity": 2,
                        "source_pdf": "input.pdf",
                        "source_page": 1,
                        "corner_reliefs": [
                            {
                                "corner": "bottom_left",
                                "relief_type": "chamfer",
                                "radius": 10,
                                "x_offset": 10,
                                "y_offset": 10,
                            }
                        ],
                        "polygon_vertices": [
                            {"x": 0, "y": 10},
                            {"x": 10, "y": 0},
                            {"x": 156.5, "y": 0},
                            {"x": 156.5, "y": 145},
                            {"x": 120, "y": 175},
                            {"x": 0, "y": 175},
                        ],
                    }
                ],
            },
            indent=2,
        ),
        encoding="utf-8",
    )

    result = run_job(job_dir, output_dir, autocad_live_policy="off")

    assert result.ok is True
    qc = json.loads((output_dir / "4042" / "4042_qc.json").read_text(encoding="utf-8"))
    assert qc["ok"] is True
    assert qc["dxf"]["polygon_corner_relief_count"] == 2
    assert qc["dxf"]["expected_corner_reliefs"] == 1
    assert qc["dxf"]["has_polygon_vertices"] is True
    assert qc["dxf"]["expected_polygon_vertex_count"] == 6
    assert qc["dxf"]["outer_contour_vertex_count"] == 6
    assert qc["plate_spec"]["polygon_vertices"][0] == {"x": 0.0, "y": 10.0}


def _sample_plate() -> PlateSpec:
    return PlateSpec(
        poz_no="P100",
        width=200,
        height=100,
        thickness=10,
        material="S355",
        quantity=2,
        holes=[
            HoleSpec(x=50, y=25, diameter=18),
            HoleSpec(x=150, y=25, diameter=18),
        ],
        confidence=0.95,
    )


def _write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = sorted({key for row in rows for key in row})
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_vector_pdf(path: Path, text_lines: list[str]) -> None:
    text_commands = "\n".join(f"({line}) Tj" for line in text_lines)
    stream = (
        "q\n"
        "0 0 m 200 0 l 200 100 l 0 100 l h S\n"
        "BT\n"
        "/F1 12 Tf\n"
        "10 10 Td\n"
        f"{text_commands}\n"
        "ET\n"
        "Q\n"
    )
    pdf = (
        "%PDF-1.4\n"
        "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n"
        "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n"
        "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 300 200] /Contents 4 0 R >> endobj\n"
        f"4 0 obj << /Length {len(stream.encode('latin-1'))} >>\n"
        "stream\n"
        f"{stream}"
        "endstream\n"
        "endobj\n"
        "trailer << /Root 1 0 R >>\n"
        "%%EOF\n"
    )
    path.write_bytes(pdf.encode("latin-1"))


def _write_multi_page_vector_pdf(path: Path, pages: list[list[str]]) -> None:
    objects = ["1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj"]
    kids = " ".join(f"{3 + index * 2} 0 R" for index in range(len(pages)))
    objects.append(f"2 0 obj << /Type /Pages /Kids [{kids}] /Count {len(pages)} >> endobj")
    for index, text_lines in enumerate(pages):
        page_obj = 3 + index * 2
        content_obj = page_obj + 1
        text_commands = "\n".join(f"({line}) Tj T*" for line in text_lines)
        stream = (
            "q\n"
            "0 0 m 200 0 l 200 100 l 0 100 l h S\n"
            "BT\n"
            "/F1 12 Tf\n"
            "10 10 Td\n"
            f"{text_commands}\n"
            "ET\n"
            "Q\n"
        )
        objects.append(
            f"{page_obj} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 300 200] /Contents {content_obj} 0 R >> endobj"
        )
        objects.append(
            f"{content_obj} 0 obj << /Length {len(stream.encode('latin-1'))} >>\n"
            "stream\n"
            f"{stream}"
            "endstream\n"
            "endobj"
        )
    pdf = "%PDF-1.4\n" + "\n".join(objects) + "\ntrailer << /Root 1 0 R >>\n%%EOF\n"
    path.write_bytes(pdf.encode("latin-1"))


def _write_vector_only_pdf(path: Path, page_count: int = 1) -> None:
    objects = ["1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj"]
    kids = " ".join(f"{3 + index * 2} 0 R" for index in range(page_count))
    objects.append(f"2 0 obj << /Type /Pages /Kids [{kids}] /Count {page_count} >> endobj")
    for index in range(page_count):
        page_obj = 3 + index * 2
        content_obj = page_obj + 1
        stream = (
            "q\n"
            f"{index} {index} m 200 {index} l 200 100 l 0 100 l h S\n"
            "20 20 m 180 20 l 180 80 l 20 80 l h S\n"
            "Q\n"
        )
        objects.append(
            f"{page_obj} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 300 200] /Contents {content_obj} 0 R >> endobj"
        )
        objects.append(
            f"{content_obj} 0 obj << /Length {len(stream.encode('latin-1'))} >>\n"
            "stream\n"
            f"{stream}"
            "endstream\n"
            "endobj"
        )
    pdf = "%PDF-1.4\n" + "\n".join(objects) + "\ntrailer << /Root 1 0 R >>\n%%EOF\n"
    path.write_bytes(pdf.encode("latin-1"))


def _skill_paths_from_markdown(text: str) -> list[str]:
    paths = []
    for line in text.splitlines():
        if "skills/" not in line:
            continue
        for part in line.split("`")[1::2]:
            if "skills/" in part and part.endswith(".md"):
                paths.append(part)
    return paths
