from __future__ import annotations

import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from .audit import get_audit_logger
from .config import RuntimePaths, ensure_autocad_import_path
from .job_fsm import JobState, get_fsm
from .manager_memory import get_manager_memory
from .memory_bridge import compute_pdf_fingerprint, get_memory_bridge
from .metrics import record_job_status
from .session_store import load_session, save_session


DEFAULT_MANAGER_SESSION_ID = "agent:teknik-ofis-muduru:dashboard"
SKILL_TARGETS = {
    "partlist": "agents/_shared/skills/ERT_PARTLIST_EXCEL_URETIMI.md",
    "dxf_geometry": "agents/_shared/skills/DXF_2013_URETIMI.md",
    "qc": "agents/_shared/skills/CIZIM_NC_KALITE_KONTROLU.md",
    "memory_bridge": "agents/_shared/skills/OGRENME_VE_HAFIZA_YONETIMI.md",
    "teknik-ofis-muduru": "agents/_shared/skills/OGRENME_VE_HAFIZA_YONETIMI.md",
}


def append_job_event(paths: RuntimePaths, job_id: str, event_type: str, payload: dict[str, Any]) -> None:
    output_dir = paths.jobs_output_root / job_id
    output_dir.mkdir(parents=True, exist_ok=True)
    event = {"type": event_type, "timestamp": _now_ms(), "payload": payload}
    with (output_dir / "events.jsonl").open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False) + "\n")


def complete_approved_job(
    paths: RuntimePaths,
    job_id: str,
    summary: dict[str, Any],
    *,
    approved_count: int | None = None,
    session_id: str | None = None,
) -> dict[str, Any]:
    """Run QC closure, partlist, retrospective, memory, and final manager notice."""
    output_dir = paths.jobs_output_root / job_id
    job_dir = paths.jobs_import_root / job_id
    fsm = get_fsm(paths.jobs_output_root)
    resolved_approved_count = _approved_plate_count(paths, job_id) if approved_count is None else approved_count

    manual_reviews = summary.get("manual_reviews") if isinstance(summary.get("manual_reviews"), list) else []
    produced = summary.get("produced") if isinstance(summary.get("produced"), list) else []
    qc_ok = summary.get("ok") is True and not manual_reviews
    append_job_event(
        paths,
        job_id,
        "qc_completed",
        {
            "ok": qc_ok,
            "produced_count": len(produced),
            "manual_review_count": len(manual_reviews),
        },
    )

    if not qc_ok:
        fsm.force_transition(job_id, JobState.AWAITING_APPROVAL, reason="qc_manual_review_pending")
        message = _blocked_message(job_id, "QC manuel inceleme bekliyor", summary=summary)
        _save_manager_notification(paths, job_id, session_id, message)
        append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message})
        append_job_event(paths, job_id, "completed", {"status": "needs_manager_approval", "stage": "qc"})
        record_job_status("awaiting_approval")
        return {"ok": False, "error": "qc_manual_review_pending", "message": message, "summary": summary}

    open_notes = _actionable_open_manager_notes(paths, job_id)
    if open_notes:
        retrospective = _write_retrospective(
            paths,
            job_id,
            summary=summary,
            partlist=_partlist_dict_from_output(output_dir) or {"ok": False, "path": None, "rows": 0, "manual_reviews": []},
            approved_count=resolved_approved_count,
            status="awaiting_approval",
            open_notes=open_notes,
            source="completion_blocked",
        )
        append_job_event(paths, job_id, "retrospective_written", {"path": retrospective.get("path"), "skill_proposal_path": retrospective.get("skill_proposal_path")})
        fsm.force_transition(job_id, JobState.AWAITING_APPROVAL, reason="open_manager_notes_pending")
        message = _blocked_message(job_id, "Acik mudur notu var", summary=summary, open_notes=open_notes)
        _save_manager_notification(paths, job_id, session_id, message)
        vault_path = _append_vault_retrospective(paths, job_id, _read_json(output_dir / "retrospective.json") or {})
        if vault_path is not None:
            retrospective["vault_path"] = _relative(vault_path, paths)
        append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message})
        append_job_event(paths, job_id, "completed", {"status": "needs_manager_approval", "stage": "manager_notes"})
        record_job_status("awaiting_approval")
        return {
            "ok": False,
            "error": "open_manager_notes_pending",
            "message": message,
            "summary": summary,
            "retrospective": retrospective,
        }

    fsm.force_transition(job_id, JobState.QC_CHECKING, reason="qc_completed")
    append_job_event(paths, job_id, "partlist_started", {"message": "Partlist generation started"})
    ensure_autocad_import_path(paths)
    from autocad_mcp.technical_office.partlist import create_partlist

    partlist = create_partlist(job_dir, output_dir)
    partlist_dict = partlist.to_dict()
    append_job_event(paths, job_id, "partlist_completed", {"ok": partlist.ok, "partlist": partlist_dict})
    get_audit_logger(paths.workspace_root).log_partlist_created(
        job_id,
        row_count=partlist_dict.get("rows", 0) if isinstance(partlist_dict, dict) else 0,
        ok=partlist.ok,
    )

    if not partlist.ok:
        fsm.force_transition(job_id, JobState.AWAITING_APPROVAL, reason="partlist_failed")
        message = _blocked_message(job_id, "Partlist uretilemedi", partlist=partlist_dict)
        _save_manager_notification(paths, job_id, session_id, message)
        append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message})
        append_job_event(paths, job_id, "completed", {"status": "needs_manager_approval", "stage": "partlist"})
        record_job_status("awaiting_approval")
        return {
            "ok": False,
            "error": "partlist_failed",
            "message": message,
            "summary": summary,
            "partlist": partlist_dict,
        }

    retrospective = _write_retrospective(
        paths,
        job_id,
        summary=summary,
        partlist=partlist_dict,
        approved_count=resolved_approved_count,
        status="completed",
    )
    append_job_event(
        paths,
        job_id,
        "retrospective_written",
        {"path": retrospective.get("path"), "skill_proposal_path": retrospective.get("skill_proposal_path")},
    )
    memory_result = _record_memory_bridge(paths, job_id)
    message = _final_message(job_id, summary=summary, partlist=partlist_dict, output_dir=output_dir)
    _save_manager_notification(paths, job_id, session_id, message)
    retrospective_payload = _read_json(output_dir / "retrospective.json") or {}
    vault_path = _append_vault_retrospective(paths, job_id, retrospective_payload)
    if vault_path is not None:
        retrospective["vault_path"] = _relative(vault_path, paths)
    append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message})

    fsm.force_transition(job_id, JobState.COMPLETED, reason="partlist_and_retrospective_complete")
    append_job_event(
        paths,
        job_id,
        "completed",
        {
            "status": "completed",
            "approved_count": resolved_approved_count,
            "partlist": partlist_dict,
            "retrospective_path": _relative(retrospective["path"], paths),
        },
    )
    record_job_status("completed")
    record_job_status("partlist_ok")
    return {
        "ok": True,
        "message": message,
        "summary": summary,
        "partlist": partlist_dict,
        "retrospective": retrospective,
        "memory_bridge": memory_result,
    }


def notify_job_blocked(
    paths: RuntimePaths,
    job_id: str,
    *,
    reason: str,
    session_id: str | None = None,
    summary: dict[str, Any] | None = None,
    partlist: dict[str, Any] | None = None,
    state: JobState = JobState.AWAITING_APPROVAL,
) -> dict[str, Any]:
    fsm = get_fsm(paths.jobs_output_root)
    fsm.force_transition(job_id, state, reason=reason)
    message = _blocked_message(job_id, reason, summary=summary, partlist=partlist, fsm_state=state.value)
    _save_manager_notification(paths, job_id, session_id, message)
    append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message})
    terminal_status = "failed" if state == JobState.FAILED else "needs_manager_approval"
    terminal_type = "failed" if state == JobState.FAILED else "completed"
    append_job_event(paths, job_id, terminal_type, {"status": terminal_status, "reason": reason})
    return {"ok": False, "error": reason, "message": message}


def learning_health(paths: RuntimePaths) -> dict[str, Any]:
    jobs_missing_retrospective: list[dict[str, Any]] = []
    jobs_with_open_notes: list[dict[str, Any]] = []
    open_note_state_violations: list[dict[str, Any]] = []
    recent_closure_events: list[dict[str, Any]] = []

    for record in _job_records(paths):
        job_id = record["job_id"]
        output_dir = paths.jobs_output_root / job_id
        summary = _read_json(output_dir / "job_summary.json") or {}
        fsm_state = get_fsm(paths.jobs_output_root).get_state(job_id).value
        partlist = _partlist_dict_from_output(output_dir)
        if _job_needs_retrospective(summary, partlist, fsm_state) and not (output_dir / "retrospective.json").exists():
            jobs_missing_retrospective.append(
                {
                    "job_id": job_id,
                    "fsm_state": fsm_state,
                    "summary_ok": summary.get("ok") if isinstance(summary, dict) else None,
                    "partlist_ok": bool(partlist and partlist.get("ok")),
                }
            )
        open_notes = _open_manager_issue_notes(output_dir)
        actionable_notes = _actionable_open_manager_notes(paths, job_id)
        if open_notes:
            item = {
                "job_id": job_id,
                "fsm_state": fsm_state,
                "open_note_count": len(open_notes),
                "affected_pozs": _affected_pozs_from_notes(open_notes),
                "actionable_note_count": len(actionable_notes),
            }
            jobs_with_open_notes.append(item)
            if actionable_notes and fsm_state != JobState.AWAITING_APPROVAL.value:
                open_note_state_violations.append(item)
        recent_closure_events.extend(_recent_closure_events(paths, job_id, limit=4))

    agent_registry = _agent_registry_health(paths)
    pending_skill_proposals = _pending_skill_proposals(paths)
    memory_patterns = get_memory_bridge(paths.workspace_root).get_pattern_stats()
    recent_closure_events.sort(key=lambda item: str(item.get("timestamp") or ""), reverse=True)
    ok = not jobs_missing_retrospective and not open_note_state_violations and bool(agent_registry.get("ok"))
    return {
        "ok": ok,
        "jobs_missing_retrospective": jobs_missing_retrospective,
        "jobs_with_open_notes": jobs_with_open_notes,
        "open_note_state_violations": open_note_state_violations,
        "memory_patterns": memory_patterns,
        "pending_skill_proposals": pending_skill_proposals,
        "agent_registry_ok": bool(agent_registry.get("ok")),
        "agent_registry": agent_registry,
        "recent_closure_events": recent_closure_events[:12],
    }


def backfill_job_learning(
    paths: RuntimePaths,
    job_id: str,
    *,
    dry_run: bool = False,
    session_id: str | None = None,
) -> dict[str, Any]:
    job_dir = paths.jobs_import_root / job_id
    output_dir = paths.jobs_output_root / job_id
    if not job_dir.exists():
        return {"ok": False, "job_id": job_id, "error": "job_not_found"}

    summary = _read_json(output_dir / "job_summary.json") or {}
    fsm_state = get_fsm(paths.jobs_output_root).get_state(job_id).value
    partlist = _partlist_dict_from_output(output_dir)
    needs_retrospective = _job_needs_retrospective(summary, partlist, fsm_state)
    open_notes = _actionable_open_manager_notes(paths, job_id)
    actions: list[str] = []

    if not needs_retrospective and not open_notes:
        return {"ok": True, "job_id": job_id, "eligible": False, "reason": "no_completed_or_partlist_state"}
    if not (output_dir / "retrospective.json").exists():
        actions.append("write_retrospective")
    actions.extend(["append_manager_vault", "record_memory_bridge", "write_skill_proposal", "append_learning_events"])
    if (
        partlist is None
        and isinstance(summary, dict)
        and summary.get("ok") is True
        and fsm_state == JobState.COMPLETED.value
        and not open_notes
    ):
        actions.insert(0, "create_partlist")

    if dry_run:
        return {
            "ok": True,
            "job_id": job_id,
            "eligible": True,
            "dry_run": True,
            "actions": list(dict.fromkeys(actions)),
            "has_retrospective": (output_dir / "retrospective.json").exists(),
            "partlist_ok": bool(partlist and partlist.get("ok")),
            "open_note_count": len(open_notes),
        }

    partlist_blocked = False
    if (
        partlist is None
        and isinstance(summary, dict)
        and summary.get("ok") is True
        and fsm_state == JobState.COMPLETED.value
        and not open_notes
    ):
        ensure_autocad_import_path(paths)
        from autocad_mcp.technical_office.partlist import create_partlist

        append_job_event(paths, job_id, "partlist_started", {"message": "Backfill partlist generation started"})
        try:
            result = create_partlist(job_dir, output_dir)
        except Exception as exc:
            partlist_blocked = True
            partlist = {
                "ok": False,
                "path": None,
                "rows": 0,
                "manual_reviews": [{"reason": "partlist_backfill_failed", "detail": str(exc)}],
            }
            append_job_event(paths, job_id, "partlist_completed", {"ok": False, "partlist": partlist})
        else:
            partlist = result.to_dict()
            append_job_event(paths, job_id, "partlist_completed", {"ok": result.ok, "partlist": partlist})
            get_audit_logger(paths.workspace_root).log_partlist_created(
                job_id,
                row_count=partlist.get("rows", 0) if isinstance(partlist, dict) else 0,
                ok=result.ok,
            )
            partlist_blocked = not result.ok

    if partlist is None:
        partlist = {"ok": False, "path": None, "rows": 0, "manual_reviews": []}

    if not partlist.get("ok"):
        partlist_blocked = True

    status = "awaiting_approval" if open_notes or partlist_blocked else "completed"
    retrospective = _write_retrospective(
        paths,
        job_id,
        summary=summary if isinstance(summary, dict) else {},
        partlist=partlist,
        approved_count=_approved_plate_count(paths, job_id),
        status=status,
        open_notes=open_notes,
        source="backfill",
    )
    append_job_event(paths, job_id, "retrospective_written", {"path": retrospective.get("path"), "skill_proposal_path": retrospective.get("skill_proposal_path"), "source": "backfill"})
    memory_result = _record_memory_bridge(paths, job_id)
    vault_path = _append_vault_retrospective(paths, job_id, _read_json(output_dir / "retrospective.json") or {})
    if vault_path is not None:
        retrospective["vault_path"] = _relative(vault_path, paths)

    fsm = get_fsm(paths.jobs_output_root)
    if open_notes or partlist_blocked:
        reason = "learning_backfill_open_notes" if open_notes else "learning_backfill_partlist_failed"
        message_reason = "Backfill acik mudur notu buldu" if open_notes else "Backfill partlist uretilemedi"
        fsm.force_transition(job_id, JobState.AWAITING_APPROVAL, reason=reason)
        message = _blocked_message(job_id, message_reason, summary=summary if isinstance(summary, dict) else {}, partlist=partlist, open_notes=open_notes)
        terminal_payload = {
            "status": "needs_manager_approval",
            "stage": "learning_backfill",
            "open_note_count": len(open_notes),
            "partlist_ok": bool(partlist.get("ok")),
        }
    else:
        fsm.force_transition(job_id, JobState.COMPLETED, reason="learning_backfill_complete")
        message = _final_message(job_id, summary=summary if isinstance(summary, dict) else {}, partlist=partlist, output_dir=output_dir)
        terminal_payload = {"status": "completed", "stage": "learning_backfill", "retrospective_path": retrospective.get("path")}

    _save_manager_notification(paths, job_id, session_id, message)
    append_job_event(paths, job_id, "manager_notification", {"session_id": _session_id(session_id), "message": message, "source": "backfill"})
    append_job_event(paths, job_id, "completed", terminal_payload)
    return {
        "ok": not bool(open_notes or partlist_blocked),
        "job_id": job_id,
        "eligible": True,
        "dry_run": False,
        "actions": list(dict.fromkeys(actions)),
        "status": status,
        "partlist": partlist,
        "retrospective": retrospective,
        "memory_bridge": memory_result,
        "message": message,
        "open_note_count": len(open_notes),
    }


def backfill_all_learning(paths: RuntimePaths, *, dry_run: bool = True) -> dict[str, Any]:
    results = [
        backfill_job_learning(paths, record["job_id"], dry_run=dry_run)
        for record in _job_records(paths)
    ]
    relevant = [item for item in results if item.get("eligible") or not item.get("ok")]
    return {
        "ok": all(bool(item.get("ok")) for item in relevant) if not dry_run else True,
        "dry_run": dry_run,
        "jobs_considered": len(results),
        "jobs_returned": len(relevant),
        "results": relevant,
    }


def _write_retrospective(
    paths: RuntimePaths,
    job_id: str,
    *,
    summary: dict[str, Any],
    partlist: dict[str, Any],
    approved_count: int,
    status: str,
    open_notes: list[dict[str, Any]] | None = None,
    source: str = "completion",
) -> dict[str, Any]:
    output_dir = paths.jobs_output_root / job_id
    produced = summary.get("produced") if isinstance(summary.get("produced"), list) else []
    manual_reviews = summary.get("manual_reviews") if isinstance(summary.get("manual_reviews"), list) else []
    calculated_metrics = _calculated_metric_details(output_dir)
    open_notes = open_notes or []
    learning = _learning_items(
        calculated_metrics=calculated_metrics,
        partlist=partlist,
        open_notes=open_notes,
        approved_count=approved_count,
    )
    payload: dict[str, Any] = {
        "job_id": job_id,
        "created_at": _now_iso(),
        "status": status,
        "source": source,
        "approved_candidate_count": approved_count,
        "produced_count": len(produced),
        "manual_review_count": len(manual_reviews),
        "qc_ok": summary.get("ok") is True,
        "partlist": partlist,
        "calculated_metrics": calculated_metrics,
        "resolved_blockages": _resolved_blockages(paths, job_id),
        "open_manager_notes": open_notes,
        "skill_promotion_policy": "proposal_pending_manual_approval",
        "learning": learning,
    }
    path = output_dir / "retrospective.json"
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    proposal_path = _write_skill_proposal(paths, job_id, payload)
    return {
        "path": _relative(path, paths),
        "vault_path": None,
        "skill_proposal_path": _relative(proposal_path, paths) if proposal_path else None,
        "metrics_calculated_count": len(calculated_metrics),
    }


def _record_memory_bridge(paths: RuntimePaths, job_id: str) -> dict[str, Any]:
    approved_path = paths.jobs_import_root / job_id / "approved_plate_specs.json"
    if not approved_path.exists():
        return {"ok": True, "skipped": True, "reason": "no_approved_specs"}
    data = _read_json(approved_path) or {}
    plates = data.get("plates") if isinstance(data.get("plates"), list) else []
    if not plates:
        return {"ok": True, "skipped": True, "reason": "empty_plates"}
    pdf_files = sorted((paths.jobs_import_root / job_id).glob("*.pdf"))
    if not pdf_files:
        return {"ok": True, "skipped": True, "reason": "no_pdf_files"}
    confidences = [_confidence(row.get("confidence", 0.5)) for row in plates if isinstance(row, dict)]
    confidence = sum(confidences) / max(len(confidences), 1)
    fingerprint = compute_pdf_fingerprint(pdf_files[0])
    get_memory_bridge(paths.workspace_root).record_extraction(
        pdf_fingerprint=fingerprint,
        job_id=job_id,
        plate_specs=plates,
        confidence=round(confidence, 3),
        source="manager_approved_completion",
    )
    return {
        "ok": True,
        "fingerprint": fingerprint[:12],
        "plate_count": len(plates),
        "confidence": round(confidence, 3),
    }


def _save_manager_notification(paths: RuntimePaths, job_id: str, session_id: str | None, message: str) -> None:
    resolved_session_id = _session_id(session_id)
    history = load_session(paths.sessions_root, resolved_session_id)
    save_session(
        paths.sessions_root,
        resolved_session_id,
        history + [{"role": "assistant", "content": message, "timestamp": _now_ms()}],
    )
    if resolved_session_id.startswith("agent:teknik-ofis-muduru:"):
        try:
            get_manager_memory(paths.workspace_root).record_turn(
                session_id=resolved_session_id,
                user_text=f"[system] {job_id} otomatik kapanis bildirimi",
                assistant_text=message,
                selected_job_id=job_id,
                history=history,
            )
        except Exception:
            pass


def _final_message(job_id: str, *, summary: dict[str, Any], partlist: dict[str, Any], output_dir: Path) -> str:
    produced = summary.get("produced") if isinstance(summary.get("produced"), list) else []
    partlist_path = partlist.get("path") or "partlist yok"
    dxf_count = len(list(output_dir.rglob("*.dxf")))
    nc1_count = len(list(output_dir.rglob("*.nc1")))
    return (
        f"`{job_id}` isi tamamlandi.\n"
        f"- Uretilen poz: {len(produced)}\n"
        "- QC: ok\n"
        f"- Partlist: {partlist_path}\n"
        f"- Ciktilar: {dxf_count} DXF, {nc1_count} NC1; dashboard'dan indirilebilir."
    )


def _blocked_message(
    job_id: str,
    reason: str,
    *,
    summary: dict[str, Any] | None = None,
    partlist: dict[str, Any] | None = None,
    open_notes: list[dict[str, Any]] | None = None,
    fsm_state: str = "awaiting_approval",
) -> str:
    lines = [f"`{job_id}` isi otomatik kapanmadi.", f"- Sebep: {reason}", f"- FSM: {fsm_state}"]
    manual_reviews = summary.get("manual_reviews") if isinstance(summary, dict) and isinstance(summary.get("manual_reviews"), list) else []
    if manual_reviews:
        lines.append("- Manuel inceleme:")
        for item in manual_reviews[:8]:
            poz = item.get("poz_no") or "poz bilinmiyor"
            page = item.get("page") or item.get("source_page") or "?"
            detail = item.get("detail") or item.get("reason") or "detay yok"
            lines.append(f"  - Poz {poz}, sayfa {page}: {detail}")
        if len(manual_reviews) > 8:
            lines.append(f"  - ... {len(manual_reviews) - 8} not daha var")
    reviews = partlist.get("manual_reviews") if isinstance(partlist, dict) and isinstance(partlist.get("manual_reviews"), list) else []
    if reviews:
        lines.append("- Partlist blokaji:")
        for item in reviews[:8]:
            poz = item.get("poz_no") or "poz bilinmiyor"
            detail = item.get("detail") or item.get("reason") or "detay yok"
            lines.append(f"  - Poz {poz}: {detail}")
    if open_notes:
        lines.append("- Acik mudur notu:")
        for note in open_notes[:8]:
            affected = note.get("affected_pozs") if isinstance(note.get("affected_pozs"), list) else []
            poz_text = ", ".join(str(value) for value in affected) if affected else "poz belirtilmemis"
            detail = str(note.get("message") or note.get("reason") or "detay yok")
            lines.append(f"  - {poz_text}: {_shorten(detail, 180)}")
    lines.append("Karar: Blokaj cozulmeden teslim/partlist kapatilmayacak.")
    return "\n".join(lines)


def _calculated_metric_details(output_dir: Path) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    for qc_path in sorted(output_dir.rglob("*_qc.json")):
        data = _read_json(qc_path) or {}
        spec = data.get("plate_spec") if isinstance(data.get("plate_spec"), dict) else {}
        if spec.get("partlist_metrics_source") != "calculated_geometry":
            continue
        details.append(
            {
                "poz_no": spec.get("poz_no") or data.get("poz_no"),
                "fields": spec.get("partlist_metrics_calculated") or [],
                "unit_surface_area_m2": spec.get("unit_surface_area_m2"),
                "unit_weight_kg": spec.get("unit_weight_kg"),
                "source": "calculated_geometry",
            }
        )
    return details


def _resolved_blockages(paths: RuntimePaths, job_id: str) -> list[dict[str, Any]]:
    approved = _read_json(paths.jobs_import_root / job_id / "approved_plate_specs.json") or {}
    plates = approved.get("plates") if isinstance(approved.get("plates"), list) else []
    corner_relief_pozs = [
        str(row.get("poz_no"))
        for row in plates
        if isinstance(row, dict) and isinstance(row.get("corner_reliefs"), list) and row.get("corner_reliefs")
    ]
    if not corner_relief_pozs:
        return []
    return [
        {
            "type": "corner_reliefs_confirmed",
            "count": len(corner_relief_pozs),
            "poz_nos": corner_relief_pozs,
        }
    ]


def _learning_items(
    *,
    calculated_metrics: list[dict[str, Any]],
    partlist: dict[str, Any],
    open_notes: list[dict[str, Any]],
    approved_count: int,
) -> list[dict[str, Any]]:
    items = [
        {
            "agent_or_skill": "teknik-ofis-muduru",
            "target_skill": SKILL_TARGETS["teknik-ofis-muduru"],
            "promotion_status": "pending_approval",
            "proposal": "Onay sonrasi zinciri partlist ve final bildirim tamamlanmadan tamamlandi sayma.",
        }
    ]
    if calculated_metrics:
        items.append(
            {
                "agent_or_skill": "partlist",
                "target_skill": SKILL_TARGETS["partlist"],
                "promotion_status": "pending_approval",
                "proposal": "PDF metrikleri yoksa plaka alan ve agirligini geometri + 7850 kg/m3 yogunlukla hesapla.",
                "affected_poz_count": len(calculated_metrics),
            }
        )
    if open_notes:
        items.append(
            {
                "agent_or_skill": "dxf_geometry",
                "target_skill": SKILL_TARGETS["dxf_geometry"],
                "promotion_status": "pending_approval",
                "proposal": "Acik mudur geometri notu varken teslim kapanisini blokla; nottaki poz/pah bilgisini yeniden aday geometriye tasima.",
                "affected_poz_count": len(_affected_pozs_from_notes(open_notes)),
            }
        )
    if partlist.get("ok") and approved_count > 0:
        items.append(
            {
                "agent_or_skill": "memory_bridge",
                "target_skill": SKILL_TARGETS["memory_bridge"],
                "promotion_status": "pending_approval",
                "proposal": "Basarili onayli plate spec'leri is kapanisinda hafizaya al.",
            }
        )
    return items


def _append_vault_retrospective(paths: RuntimePaths, job_id: str, payload: dict[str, Any]) -> Path | None:
    path = paths.workspace_root / "manager_vault" / "jobs" / f"{_safe_filename(job_id)}.md"
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "",
        f"## Retrospective - {_now_iso()}",
        f"- Status: {payload.get('status')}",
        f"- Approved candidates: {payload.get('approved_candidate_count')}",
        f"- Produced positions: {payload.get('produced_count')}",
        f"- QC ok: {payload.get('qc_ok')}",
        f"- Partlist rows: {(payload.get('partlist') or {}).get('rows')}",
        f"- Calculated metrics: {len(payload.get('calculated_metrics') or [])}",
    ]
    try:
        with path.open("a", encoding="utf-8") as handle:
            handle.write("\n".join(lines) + "\n")
    except OSError:
        return None
    return path


def _write_skill_proposal(paths: RuntimePaths, job_id: str, payload: dict[str, Any]) -> Path | None:
    proposals = payload.get("learning") if isinstance(payload.get("learning"), list) else []
    if not proposals:
        return None
    root = paths.suite_root / "journal" / "skill_proposals"
    root.mkdir(parents=True, exist_ok=True)
    path = root / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_safe_filename(job_id)}_auto_close.md"
    lines = [
        f"# Skill Proposal - {job_id}",
        "",
        f"Created: {_now_iso()}",
        "",
        "These are proposals only. Agent and skill source files were not changed automatically.",
        "",
        "## Reusable Learnings",
    ]
    for item in proposals:
        lines.append(
            f"- [{item.get('promotion_status') or 'pending_approval'}] "
            f"{item.get('agent_or_skill')} -> {item.get('target_skill')}: {item.get('proposal')}"
        )
    lines.extend(
        [
            "",
            "## Evidence",
            f"- Produced positions: {payload.get('produced_count')}",
            f"- Partlist rows: {(payload.get('partlist') or {}).get('rows')}",
            f"- Calculated metrics: {len(payload.get('calculated_metrics') or [])}",
        ]
    )
    try:
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except OSError:
        return None
    return path


def _read_json(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _relative(path: Path | str | None, paths: RuntimePaths) -> str | None:
    if path is None:
        return None
    candidate = Path(path)
    try:
        return candidate.resolve().relative_to(paths.suite_root.resolve()).as_posix()
    except Exception:
        return str(path)


def _session_id(session_id: str | None) -> str:
    value = str(session_id or "").strip()
    return value or DEFAULT_MANAGER_SESSION_ID


def _safe_filename(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_.-]+", "_", value).strip("._") or "job"


def _approved_plate_count(paths: RuntimePaths, job_id: str) -> int:
    approved = _read_json(paths.jobs_import_root / job_id / "approved_plate_specs.json") or {}
    plates = approved.get("plates") if isinstance(approved, dict) else None
    return len(plates) if isinstance(plates, list) else 0


def _partlist_dict_from_output(output_dir: Path) -> dict[str, Any] | None:
    review = _read_json(output_dir / "partlist_manual_review_required.json")
    workbooks = sorted(output_dir.glob("*_partlist.xlsx"))
    if workbooks:
        workbook = workbooks[-1]
        return {"ok": True, "path": str(workbook), "rows": _partlist_row_count(workbook), "manual_reviews": []}
    if review is not None:
        return {"ok": False, "path": None, "rows": 0, "manual_reviews": review}
    return None


def _partlist_row_count(path: Path) -> int:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            return max(0, int(sheet.max_row or 1) - 1)
        finally:
            workbook.close()
    except Exception:
        return 0


def _job_records(paths: RuntimePaths) -> list[dict[str, str]]:
    if not paths.jobs_import_root.exists():
        return []
    return [{"job_id": path.name} for path in sorted(paths.jobs_import_root.iterdir()) if path.is_dir()]


def _job_needs_retrospective(summary: Any, partlist: dict[str, Any] | None, fsm_state: str) -> bool:
    return partlist is not None or fsm_state == JobState.COMPLETED.value


def _open_manager_issue_notes(output_dir: Path) -> list[dict[str, Any]]:
    path = output_dir / "manager_issue_notes.jsonl"
    if not path.exists():
        return []
    notes: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        try:
            note = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(note, dict) and note.get("status") in {None, "open"}:
            notes.append(note)
    return notes


def _actionable_open_manager_notes(paths: RuntimePaths, job_id: str) -> list[dict[str, Any]]:
    output_dir = paths.jobs_output_root / job_id
    notes = _open_manager_issue_notes(output_dir)
    if not notes:
        return []
    approved = _read_json(paths.jobs_import_root / job_id / "approved_plate_specs.json") or {}
    plates = approved.get("plates") if isinstance(approved, dict) and isinstance(approved.get("plates"), list) else []
    approved_by_poz = {
        str(item.get("poz_no")): item
        for item in plates
        if isinstance(item, dict) and item.get("poz_no")
    }
    known_pozs = set(approved_by_poz)
    actionable: list[dict[str, Any]] = []
    for note in notes:
        tags = {str(tag).lower() for tag in note.get("tags", []) if str(tag).strip()} if isinstance(note.get("tags"), list) else set()
        if not ({"pah/kose eksigi", "poligon kontur", "gorsel analiz notu"} & tags):
            continue
        if known_pozs:
            affected = [
                str(value)
                for value in note.get("affected_pozs", [])
                if str(value).strip() in known_pozs
            ] if isinstance(note.get("affected_pozs"), list) else []
        else:
            affected = [
                str(value)
                for value in note.get("affected_pozs", [])
                if re.fullmatch(r"[0-9]{3,6}", str(value).strip())
            ] if isinstance(note.get("affected_pozs"), list) else []
        if not affected:
            affected = _extract_known_pozs_from_text(str(note.get("message") or ""), known_pozs)
        if not affected:
            continue
        unresolved: list[str] = []
        for poz_no in affected:
            approved_row = approved_by_poz.get(poz_no)
            if approved_row is None:
                unresolved.append(poz_no)
                continue
            if not approved_row.get("corner_reliefs"):
                unresolved.append(poz_no)
        if unresolved:
            item = dict(note)
            item["affected_pozs"] = unresolved
            actionable.append(item)
    return actionable


def _extract_known_pozs_from_text(text: str, known_pozs: set[str]) -> list[str]:
    if not text or not known_pozs:
        return []
    candidates = list(dict.fromkeys(re.findall(r"\b[0-9]{3,6}\b", text)))
    return [value for value in candidates if value in known_pozs]


def _affected_pozs_from_notes(notes: list[dict[str, Any]]) -> list[str]:
    pozs: list[str] = []
    for note in notes:
        values = note.get("affected_pozs")
        if isinstance(values, list):
            pozs.extend(str(value) for value in values if str(value).strip())
    return list(dict.fromkeys(pozs))


def _agent_registry_health(paths: RuntimePaths) -> dict[str, Any]:
    try:
        registry = _read_json(paths.registry_path) or {}
    except Exception:
        registry = {}
    agents = registry.get("agents") if isinstance(registry, dict) else []
    missing_brains: list[str] = []
    active_count = 0
    if isinstance(agents, list):
        for agent in agents:
            if not isinstance(agent, dict) or agent.get("status") != "active":
                continue
            active_count += 1
            brain = agent.get("brain")
            if isinstance(brain, str) and not (paths.suite_root / "agents" / brain).exists():
                missing_brains.append(str(agent.get("id") or brain))
    missing_skill_targets = [
        target
        for target in sorted(set(SKILL_TARGETS.values()))
        if not (paths.suite_root / target).exists()
    ]
    return {
        "ok": active_count > 0 and not missing_brains and not missing_skill_targets,
        "active_count": active_count,
        "missing_brains": missing_brains,
        "missing_skill_targets": missing_skill_targets,
    }


def _pending_skill_proposals(paths: RuntimePaths) -> dict[str, Any]:
    root = paths.suite_root / "journal" / "skill_proposals"
    files = sorted(root.glob("*.md")) if root.exists() else []
    return {
        "count": len(files),
        "items": [
            {"path": _relative(path, paths), "status": "pending_approval"}
            for path in files[-12:]
        ],
    }


def _recent_closure_events(paths: RuntimePaths, job_id: str, *, limit: int = 4) -> list[dict[str, Any]]:
    path = paths.jobs_output_root / job_id / "events.jsonl"
    if not path.exists():
        return []
    selected = {"production_started", "qc_completed", "partlist_completed", "retrospective_written", "manager_notification", "completed", "failed"}
    events: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines()[-120:]:
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(event, dict) and event.get("type") in selected:
            events.append({"job_id": job_id, **event})
    return events[-limit:]


def _shorten(value: str, limit: int) -> str:
    compact = re.sub(r"\s+", " ", value).strip()
    if len(compact) <= limit:
        return compact
    return compact[: max(0, limit - 3)].rstrip() + "..."


def _confidence(value: Any) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return 0.5
    return parsed if 0 <= parsed <= 1 else 0.5


def _now_ms() -> int:
    return int(time.time() * 1000)


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat()
